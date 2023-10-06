from openpyxl import load_workbook
import random

# config
groupSize = 4                         # group size to generate. larger groups will still be parsed for names
group_skip = "--"                     # text to skip a group member if a group wants to be smaller than groupSize
file_names = "names.txt"              # names file. 1 name per line
file_spreadsheet = "spreadsheet.xlsx" # spreasheet file

# load names
names = []
with open(file_names, 'r') as f:
    for line in f:
        names.append(line.strip().lower())

# load spreadsheet
wb = load_workbook(file_spreadsheet)
sheet = wb.worksheets[0]

# parse all groups for names:
# removes known names from list
# and prints unknown names to console
def parse_groups():
    unknownNames = False

    # loop groups (rows)
    for i in range(2, sheet.max_row + 1):

        # loop names (collums)
        for j in range(2, sheet.max_column + 1):
            name = str(sheet.cell(row=i, column=j).value).lower()
            
            # skip cells marked to skip or empty
            if name == "none" or name == group_skip:
                continue

            if name in names:
                names.remove(name)

            else:
                unknownNames = True
                print("Invalid name in group " + str(i - 1) + ": " + str(name))

    return unknownNames

def fill_groups():

    # loop groups (rows)
    for i in range(2, sheet.max_row + len(names) + 1):

        # loop names (collums)
        for j in range(2, groupSize + 2):

            # stop when all names have been used
            if not names:
                # save spreadsheet
                wb.save(file_spreadsheet)
                return

            name = sheet.cell(row=i, column=j).value
            
            # skip cells that have names in them
            if name is not None:
                continue
            
            randName = random.choice(names)
            names.remove(randName)
            sheet.cell(row=i, column=j).value = randName

            print("Added " + randName + " to group " + str(i - 1) + " in cell " + str(j))

def main():
    colors = { "green": '\033[92m', "yellow": '\033[93m', "end": '\033[0m' }
    unknownNames = parse_groups()

    if unknownNames:
        print(colors["yellow"] + "Please correct or delete invalid names above then run script again" + colors["end"])
        print("Names without groups: " + str(names))
        return
    
    fill_groups()
    print(colors["green"] + "done!" + colors["end"])
    
if __name__ == "__main__": 
    main()